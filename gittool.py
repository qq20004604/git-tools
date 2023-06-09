import os
import gitlab
import logging
import shutil
import pathlib
import time
import yaml
from openpyxl import Workbook
from openpyxl import load_workbook
from git import Repo
import re
import datetime
import ctypes
import json


class GitTool:
    def __init__(self):
        # 配置
        self.config = None
        # gitlab api的客户端
        self.gl = None
        self.workbook = None
        # excel
        self.worksheet = None
        # 日志句柄
        self.match_logger = None
        self.branch_logger = None
        self.err_logger = None
        self.search_lib = None
        # 时间限制
        self.commit_since_before = None

    # 初始化
    def init(self):
        # 删除历史文件
        self._remove_oldfile()
        # 设置日志
        self._set_log()
        # 启动 excel
        self._open_excel()
        # 读取配置
        self._read_config()

    # 读取配置
    def _read_config(self):
        # 读取配置文件
        # with open('config_private.yml') as file:
        with open('config.yml') as file:
            config = yaml.safe_load(file)

        # 兼容性处理（开始）
        # 1. 处理 branch 提取出来
        if 'branch' not in config:
            branch_dict = {
                'branch': {
                    'branch_match_type': config['type_group']['branch_match_type'],
                    'branch_match_name': config['type_group']['branch_match_name'],
                    'branch_limit': config['type_group']['branch_limit'],
                    'commit_since_before': config['type_group']['commit_since_before'],
                }
            }
            config.update(branch_dict)

        # 兼容性处理（结束）

        self.config = config
        # 设置 GitLab API 客户端
        self.gl = gitlab.Gitlab(config['gitlab_api_url'], private_token=self.config['gitlab_api_access_token'])

        # 检查配置
        if self.config['model'] not in ['group', 'repository', 'repositories']:
            print("model 模式错误，只能是 group 群组模式，repository 单项目模式，repositories 多项目模式")
            exit()
        return config

    # 删除历史文件
    def _remove_oldfile(self):
        # 开始前先删除文件
        files_to_delete = ['./log/match.log', './log/branch.log', './log/err.log', './log/output.log',
                           './log/match.xlsx']

        for file in files_to_delete:
            if os.path.exists(file):
                os.remove(file)
                print(f"文件 {file} 已删除")
            else:
                print(f"文件 {file} 不存在，无需清理")

    # 检查excel
    def _open_excel(self):
        # 尝试打开 Excel 文件以检查是否可以读写
        try:
            self.workbook = load_workbook(filename='./log/match.xlsx')
            self.worksheet = self.workbook.active
            # 清空之前的内容
            for row in self.worksheet.iter_rows(min_row=2, min_col=1):
                for cell in row:
                    cell.value = None
        except FileNotFoundError:
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "匹配结果"
            self.worksheet.append(["仓库名", "分支名", "文件路径", "文件行数", "该行内容"])
            self.worksheet = self.workbook.active
        except IOError as e:
            logging.error(f"无法打开 Excel 文件：{e}")
            exit()

        # 创建一个新的 Excel 工作簿
        if not self.worksheet.cell(row=1, column=1).value:
            self.worksheet.title = "匹配结果"
            self.worksheet.append(["仓库名", "分支名", "文件路径", "文件行数"])

    # 日志设置
    def _set_log(self):
        if not os.path.exists("log"):
            os.mkdir("log")
            print("日志目录已创建")
        else:
            print("日志目录已存在")
        # 配置 output.log 记录器
        logging.basicConfig(filename='log/output.log', level=logging.INFO, format='%(message)s')

        # 创建一个新的日志记录器并配置 match.log
        self.match_logger = logging.getLogger('match')
        match_handler = logging.FileHandler('log/match.log')
        match_handler.setFormatter(logging.Formatter('%(message)s'))
        self.match_logger.addHandler(match_handler)
        self.match_logger.setLevel(logging.INFO)

        # 创建一个新的日志记录器并配置 branch.log
        self.branch_logger = logging.getLogger('branch')
        branch_handler = logging.FileHandler('log/branch.log')
        branch_handler.setFormatter(logging.Formatter('%(message)s'))
        self.branch_logger.addHandler(branch_handler)
        self.branch_logger.setLevel(logging.INFO)

        self.err_logger = logging.getLogger('err')
        err_handler = logging.FileHandler('log/err.log')
        err_handler.setFormatter(logging.Formatter('%(message)s'))
        self.err_logger.addHandler(err_handler)
        self.err_logger.setLevel(logging.INFO)

    # 遍历所有符合的文件，然后返回文件名
    # 入参是本地地址
    def find_files_by_match_name(self, local_repo_path, path=""):
        js_files = []

        for root, dirs, files in os.walk(os.path.join(local_repo_path, path)):
            # 排除 .git 目录和其他隐藏文件等
            dirs[:] = [d for d in dirs if not d.startswith('.')]
            files[:] = [f for f in files if not f.startswith('.')]

            # 全文件匹配
            if self.config['file_match']['type'] == "all":
                for file in files:
                    js_file_path = os.path.relpath(os.path.join(root, file), local_repo_path)
                    js_files.append(js_file_path)
            # 字符串匹配文件名
            elif self.config['file_match']['type'] == "normal":
                for file in files:
                    # 如果这些某个文件以这个为结尾
                    if any(type in file for type in self.config['file_match']['file_type']):
                        js_file_path = os.path.relpath(os.path.join(root, file), local_repo_path)
                        js_files.append(js_file_path)
            # 后缀名匹配
            elif self.config['file_match']['type'] == "ext":
                for file in files:
                    # 如果这些某个文件以这个为结尾
                    if any(file.endswith(ext) for ext in self.config['file_match']['file_type']):
                        js_file_path = os.path.relpath(os.path.join(root, file), local_repo_path)
                        js_files.append(js_file_path)
            # 正则匹配
            elif self.config['file_match']['type'] == "regexp":
                for file in files:
                    # 解析字符串为正则表达式
                    regex = re.compile(self.config['file_match']['file_type'])
                    if len(regex.findall(file)) > 0:
                        js_file_path = os.path.relpath(os.path.join(root, file), local_repo_path)
                        js_files.append(js_file_path)
            else:
                self.err_logger.info("未选择文件匹配模式，程序退出")
                exit()

        return js_files

    # 定义一个函数，用于在文件内容中搜索给定的【字符串】
    def search_string_in_file_by_python(self, file_path, search_string):
        res = {
            "MatchedLines": [],
            "Error": None,
        }

        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                # 处理文件内容
                file_content = file.read()
                for line_number, line in enumerate(file_content.split('\n')):
                    if search_string in line:
                        item = {
                            "LineNumber": line_number + 1,
                            "Line": line
                        }
                        res['MatchedLines'].append(item)
                        # print("匹配成功", line)
        except FileNotFoundError:
            err = f"文件未找到：{file_path}"
            res['Error'] = err
            print(err)
        except PermissionError:
            err = f"没有权限读取文件：{file_path}"
            res['Error'] = err
            print(err)
        except IOError as e:
            err = f"读取文件时发生错误：{e}"
            res['Error'] = err
            print(err)
        finally:
            # if len(res['MatchedLines']) > 0:
            #     print(res)
            return res

    # 入参：文件路径，搜索字符串
    # 出参：一个对象。
    #   result.get('Error') 为 True 表示有报错信息，通过 result['Error'] 获取报错信息
    #   result['MatchedLines'] 为匹配成功的数据，len=0说明没有任何被匹配成功
    #   result['MatchedLines'][0] 示例结果 {'LineNumber': 353, 'Line': '仓库名-分支名-文件相对路径-行数-该行代码内容'}
    def search_string_in_file_by_go(self, file_path, search_string):
        # 如果没有初始化 go 库，则初始化
        if self.search_lib is None:
            self.search_lib = ctypes.CDLL(os.path.abspath("./lib/search.so"))
            self.search_lib.searchStringInFileWrapper.argtypes = [ctypes.c_char_p, ctypes.c_char_p]
            self.search_lib.searchStringInFileWrapper.restype = ctypes.c_char_p
        # 将字符串转义
        c_file_path = ctypes.c_char_p(file_path.encode('utf-8'))
        c_search_string = ctypes.c_char_p(search_string.encode('utf-8'))
        # 调用go库
        c_result = self.search_lib.searchStringInFileWrapper(c_file_path, c_search_string)
        # 拉取返回结果
        result = json.loads(ctypes.string_at(c_result).decode('utf-8'))
        return result

    # 群组模式：过滤项目（将不需要处理的群组过滤掉）
    def _filter_projects_by_group(self):
        # 根据群组ID拿到所有项目
        if self.config['type_group']['group_id'] is None:
            self.err_logger.info("群组项目ID（type_group.group_id）为空")
        group = self.gl.groups.get(self.config['type_group']['group_id'])
        projects = group.projects.list(all=True)

        # 对仓库进行过滤，如果是普通模式，则遍历
        if self.config['type_group']['project_match_type'] == 'normal':
            l = []
            for project in projects:
                if self.config['type_group']['project_match_str'] in project.name:
                    l.append(project)
            projects = l
        elif self.config['type_group']['project_match_type'] == 'regexp':
            # 正则模式
            l = []
            regx = re.compile(self.config['type_group']['project_match_str'])
            for project in projects:
                if regx.findall(project.name):
                    l.append(project)
            projects = l

        # projects_name_log = [project.name for project in projects]
        msg = "本次处理的项目有：", [project.name for project in projects]
        print(msg)
        logging.info(msg)
        return projects

    # 过滤分支（将不需要处理的分支过滤掉）
    # 参数1：branches 指该项目下所有分支
    # 参数2：repo 指项目（一个对象）
    def _filter_branches(self, branches, repo):
        selected_branches = []
        # 如果匹配模式是全部分支
        if self.config['branch']['branch_match_type'] == "all":
            print("当前寻找分支模式是：全部分支")
            selected_branches = [branch.name for branch in branches]
        elif self.config['branch']['branch_match_type'] == "name_match":
            print("当前寻找分支模式是：名称匹配")
            all_branch_names = [branch.name for branch in branches]
            for branch_name in all_branch_names:
                # 解析字符串
                branch_match_name = self.config['branch']['branch_match_name']
                # 避免不写情况，不写则默认为所有
                if len(branch_match_name) == 0:
                    err_msg = "错误：当 branch.branch_match_type 为 name_match 时，branch.branch_match_name 不能可为空。"
                    print(err_msg)
                    self.err_logger.info(err_msg)
                    exit()

                regex = re.compile(branch_match_name)
                # 分支名进行匹配
                if len(regex.findall(branch_name)) > 0:
                    # 匹配成功则添加
                    selected_branches.append(branch_name)
        elif self.config['branch']['branch_match_type'] == "last_commit_time":
            # 存储每个分支的最新提交时间和分支名称的元组
            branch_commits = []
            # 按最后提交时间进行匹配
            # 遍历所有分支
            for branch in branches:
                # print(f"正在获取分支：{branch.name} 的最新提交")
                delta = datetime.timedelta(days=self.commit_since_before)
                start_date = datetime.datetime.now() - delta
                # 使用15天前的日期和当前日期作为日期范围
                end_date = datetime.datetime.now()
                # 获取该分支的最新提交，限制只返回1个
                commit_list = repo.commits.list(ref_name=branch.name, since=start_date, until=end_date, get_all=True)
                # print(f"总计有 {len(commit_list)} 个提交")
                if len(commit_list) == 0:
                    msg = f"分支：{branch.name} 最近{self.commit_since_before}天，无提交，跳过"
                    print(msg)
                    self.branch_logger.info(msg)
                    continue
                elif len(commit_list) > 20:
                    msg = f"分支：{branch.name} 最近{self.commit_since_before}天，提交超过20次"
                    print(msg)
                    self.branch_logger.info(msg)
                commit = commit_list[0]
                # 将最新提交时间和分支名称作为一个元组添加到branch_commits列表中
                branch_commits.append((commit.committed_date, branch.name))

            # 按最近提交时间对分支进行排序，并提取排好序的分支名称。b[1] 指取 branch.name
            selected_branches = [b[1] for b in sorted(branch_commits, reverse=True)]
            print(f"所有分支排序后为：{selected_branches}")
        else:
            err_msg = "错误：当 model 为 group 时，未找到合法的 branch.branch_match_type"
            print(err_msg)
            self.err_logger.info(err_msg)
            exit()

        # 此时拿到该项目下符合要求的分支，准备clone项目并进行处理
        # 非全部分支模式下，则限定分支进行处理
        if self.config['branch']['branch_match_type'] != 'all' and self.config['branch']['branch_limit'] != 'all':
            if self.config['branch']['branch_limit'] > 0:
                print("目前限定处理分支数量是：", self.config['branch']['branch_limit'])
                # 则限制指定数量的分支数
                selected_branches = selected_branches[:self.config['branch']['branch_limit']]

        return selected_branches

    # 将项目clone到本地，然后进行处理（调用函数遍历项目查询字符串）
    def _clone_and_deal(self, branch, repo):
        logging.info(f"----- 小分割线 -----")
        msg = f"正在处理分支：{branch}"
        logging.info(msg)
        print(msg)
        # 测试模式下，这里直接返回，不进行具体处理
        if self.config['test_mode'] is True and self.config['test_only_show_branch'] is True:
            msg = f"测试模式警告：test_mode 与 test_only_show_branch 为 true，因此不处理分支，直接返回"
            logging.info(msg)
            print(msg)
            return

        # 检查目录是否存在，如果存在则删除
        local_repo_path = f"tempdir/{repo.name}-{branch}"
        local_repo_pathlib = pathlib.Path(local_repo_path)
        if local_repo_pathlib.exists():
            shutil.rmtree(local_repo_path, ignore_errors=True)

        # 克隆仓库到本地临时目录
        local_repo_path = f"tempdir/{repo.name}-{branch}"
        Repo.clone_from(repo.http_url_to_repo, local_repo_path, branch=branch)
        print(f"已克隆到本地：{local_repo_path}")

        try:
            print("开始获取所有匹配的文件")
            # 从本地仓库获取文件列表
            js_files = self.find_files_by_match_name(local_repo_path)

            # 添加分支信息到 branch.log
            local_repo = Repo(local_repo_path)
            branch_commit = local_repo.commit(branch)
            commit_time = time.strftime('%Y-%m-%d %H:%M:%S', time.gmtime(branch_commit.committed_date))
            branch_info = f"{repo.name}-{branch}-{branch_commit.hexsha}-{commit_time}-{branch_commit.author}"
            self.branch_logger.info(branch_info)

            print("开始遍历所有文件进行查询")
            # 遍历所有匹配的文件
            for js_file_path in js_files:
                logging.info(f"正在处理文件：{os.path.abspath(os.path.join(local_repo_path, js_file_path))}")
                print(f"正在处理文件：{os.path.abspath(os.path.join(local_repo_path, js_file_path))}")
                # if os.path.abspath(os.path.join(local_repo_path,
                #                                 js_file_path)) == "/Users/wangdong/githubProject/python3-lib/在git项目里寻找某个接口是否有调用/Special-864/src/api/wd.js":
                #     print("debug")
                if self.config['file_search_engine'] == "python":
                    # 使用 python 来查询文件
                    search_result = self.search_string_in_file_by_python(
                        os.path.join(local_repo_path, js_file_path),
                        self.config['string_to_search'])
                else:
                    search_result = self.search_string_in_file_by_go(
                        os.path.join(local_repo_path, js_file_path),
                        self.config['string_to_search'])

                if search_result['Error'] is not None:
                    # 如果有报错信息，说明在查找该文件的时候出问题了
                    self.err_logger.info(search_result['Error'])
                else:
                    # 此时正常，拿到匹配的行数
                    matched_lines = search_result['MatchedLines']

                    # 长度 > 0 ，说明有匹配的代码
                    if len(matched_lines) > 0:
                        for matched_line in matched_lines:
                            match_info = f"{repo.name}-{branch}-{js_file_path}-{matched_line['LineNumber']}-{matched_line['Line'].strip()}"
                            logging.info(match_info)
                            # 写入 match.log
                            self.match_logger.info(match_info)

                            # 将匹配结果添加到 Excel 工作表
                            self.worksheet.append([repo.name, branch, js_file_path, matched_line['LineNumber'],
                                                   matched_line['Line'].strip()])


        except Exception as e:
            self.err_logger.error(f"项目：{repo.name}，处理分支 {branch} 时出错：{e}")
        finally:
            # 删除本地仓库临时目录
            shutil.rmtree(local_repo_path, ignore_errors=True)

    # 分支处理模式
    def _branch_option(self):
        if self.config['branch']['branch_match_type'] == 'all':
            msg = f"当前分支匹配模式是：全部分支"
            logging.info(msg)
            print(msg)
        elif self.config['branch']['branch_match_type'] == 'last_commit_time':
            # 计算n天前的日期
            self.commit_since_before = self.config['branch']['commit_since_before'] or 15
            msg = f"当前寻找分支模式是：按最后提交时间。时间限制为最近{self.commit_since_before}天提交过的分支，然后对其进行排序，按提交顺序从近到远获取若干个分支进行处理。"
            logging.info(msg)
            print(msg)
            msg = f"分支数为：{self.config['branch']['branch_limit']}"
            logging.info(msg)
            print(msg)
        elif self.config['branch']['branch_match_type'] == 'name_match':
            msg = f"当前寻找分支模式是：名称匹配。关键词为：{self.config['branch']['branch_match_name']}"
            logging.info(msg)
            print(msg)
            msg = f"分支数为：{self.config['branch']['branch_limit']}"
            logging.info(msg)
            print(msg)

    # 群组模式
    def _search_by_model_group(self):
        print("—————— 配置说明分割线（开始） ——————")
        logging.info("—————— 配置说明分割线（开始） ——————")
        msg = "本次处理的模式是：群组模式（group）"
        logging.info(msg)
        print(msg)
        msg = f"当前的群组ID是：{self.config['type_group']['group_id']}"
        logging.info(msg)
        print(msg)
        if self.config['type_group']['project_match_type'] == 'all':
            msg = "项目筛选条件是：所有项目"
        elif self.config['type_group']['project_match_type'] == 'normal':
            msg = f"项目筛选条件是：名称匹配。关键词是：{self.config['type_group']['project_match_str']}"
        elif self.config['type_group']['project_match_type'] == 'regexp':
            msg = f"项目筛选条件是：正则表达式。正则表达式是：{self.config['type_group']['project_match_str']}"
        logging.info(msg)
        print(msg)
        self._branch_option()
        print("—————— 配置说明分割线（结束） ——————")
        logging.info("—————— 配置说明分割线（结束） ——————")

        # 先获取要处理哪些项目
        projects = self._filter_projects_by_group()

        # 遍历项目
        for project in projects:
            logging.info("==== 大分割线 ====")
            logging.info(f"正在处理项目：{project.name}")
            print("==== 大分割线 ====")
            print(f"正在处理项目：{project.name}")
            # 根据项目id获取项目仓库
            repo = self.gl.projects.get(project.id)
            # 再获取到所有分支
            branches = repo.branches.list(get_all=True)

            # 过滤分支，只获取指定分支
            selected_branches = self._filter_branches(branches, repo)

            msg = f"总计{len(selected_branches)}个分支。"
            print(msg)
            logging.info(msg)
            for branch in selected_branches:
                # 将项目clone到本地进行处理
                self._clone_and_deal(branch, repo)

    # 单项目模式
    def _search_by_model_repository(self):
        print("—————— 配置说明分割线（开始） ——————")
        logging.info("—————— 配置说明分割线（开始） ——————")
        msg = "本次处理的模式是：单项目模式（repository）"
        logging.info(msg)
        print(msg)
        self._branch_option()

        # 先判断是什么模式
        if self.config['type_repository']['repository_model'] == 'name':
            repo_name = self.config['type_repository']['repository_name']
            # 处理一下，如果是全链接，则移除前面那部分gitlab的地址
            repo_name = repo_name.replace(self.config['gitlab_api_url'], '')
            # 再处理一下，如果第一个字母是 / ，则删除
            if repo_name.startswith("/") is True:
                repo_name = repo_name[1:]
            repo = self.gl.projects.get(repo_name)
            msg = f"项目选择方式是：根据名称获取项目。项目名称是：{repo_name}"
            logging.info(msg)
            print(msg)
        elif self.config['type_repository']['repository_model'] == 'id':
            repo_id = self.config['type_repository']['repository_id']
            repo = self.gl.projects.get(repo_id)
            msg = f"项目选择方式是：根据仓库ID获取项目。项目ID是：{repo_id}，项目名称是：{repo.name}"
            logging.info(msg)
            print(msg)
        else:
            self.err_logger.info("无效配置，请检查type_repository.repository_model，程序已退出")
            exit()

        print("—————— 配置说明分割线（结束） ——————")
        logging.info("—————— 配置说明分割线（结束） ——————")

        # 再获取到所有分支
        branches = repo.branches.list(get_all=True)

        # 过滤分支，只获取指定分支
        selected_branches = self._filter_branches(branches, repo)
        msg = f"总计{len(selected_branches)}个分支。"
        print(msg)
        logging.info(msg)
        for branch in selected_branches:
            # 将项目clone到本地进行处理
            self._clone_and_deal(branch, repo)

    # 多项目模式
    def _get_projects_by_model_repositories(self):
        print("—————— 配置说明分割线（开始） ——————")
        logging.info("—————— 配置说明分割线（开始） ——————")
        msg = "本次处理的模式是：多项目模式（repositories）"
        logging.info(msg)
        print(msg)
        self._branch_option()
        project_list = []

        # 先判断是什么模式
        if self.config['type_repositories']['repositories_model'] == 'name':
            for repo_name in self.config['type_repositories']['repositories_name']:
                # 处理一下，如果是全链接，则移除前面那部分gitlab的地址
                repo_name = repo_name.replace(self.config['gitlab_api_url'], '')
                # 再处理一下，如果第一个字母是 / ，则删除
                if repo_name.startswith("/") is True:
                    repo_name = repo_name[1:]
                repo = self.gl.projects.get(repo_name)
                project_list.append(repo)
            name_list = [project.name for project in project_list]
            msg = f"项目选择方式是：根据名称获取项目。项目有：{name_list}"
            logging.info(msg)
            print(msg)
        elif self.config['type_repositories']['repositories_model'] == 'id':
            for repo_id in self.config['type_repositories']['repositories_id']:
                repo = self.gl.projects.get(repo_id)
                project_list.append(repo)

            name_list = [project.name for project in project_list]
            msg = f"项目选择方式是：根据仓库ID获取项目。项目ID是：{self.config['type_repositories']['repositories_id']}，项目名称是：{name_list}"
            logging.info(msg)
            print(msg)
        else:
            self.err_logger.info("无效配置，type_repositories.repositories_model，程序已退出")
            exit()

        print("—————— 配置说明分割线（结束） ——————")
        logging.info("—————— 配置说明分割线（结束） ——————")

        # 遍历所有项目
        for project in project_list:
            logging.info("==== 大分割线 ====")
            logging.info(f"正在处理项目：{project.name}")
            print("==== 大分割线 ====")
            print(f"正在处理项目：{project.name}")
            # 再获取到所有分支
            branches = project.branches.list(get_all=True)

            # 过滤分支，只获取指定分支
            selected_branches = self._filter_branches(branches, project)
            msg = f"总计{len(selected_branches)}个分支。"
            print(msg)
            logging.info(msg)
            for branch in selected_branches:
                # 将项目clone到本地进行处理
                self._clone_and_deal(branch, project)

    # 本地模式
    def _get_projects_by_model_local(self):
        pass

    # todo 拿到项目的目录list，依次在文件列表里进行匹配

    # 主执行程序
    def run(self):
        # 添加本次任务开始执行时间
        logging.info(f"本次任务开始执行时间: {time.strftime('%Y-%m-%d %H:%M:%S')}")

        # 调用各自模式的函数去处理
        if self.config['model'] == 'group':
            self._search_by_model_group()
        elif self.config['model'] == 'repository':
            self._search_by_model_repository()
        elif self.config['model'] == 'repositories':
            self._get_projects_by_model_repositories()
        elif self.config['model'] == 'local':
            pass
        else:
            err_msg = "未选择任何一个合法模式，请检查配置文件"
            print(err_msg)
            self.err_logger.info(err_msg)
            exit()
        shutil.rmtree('tempdir', ignore_errors=True)
        self.workbook.save('./log/match.xlsx')
        print("处理完毕")


if __name__ == '__main__':
    gt = GitTool()
    # 删除历史文件
    gt._remove_oldfile()
    # 设置日志
    gt._set_log()
    # 启动 excel
    gt._open_excel()
    # 读取配置
    gt._read_config()
    # 启动程序
    gt.run()

    # with open('config.yml', 'r') as file:
    #     data = yaml.safe_load(file)
    #     print(data['file_type'])
    #
    # start_time = time.time()
    #
    # file_path = "./output.log"
    # search_string = "approvalComments/"
    #
    # # hello_lib = ctypes.CDLL(os.path.abspath("./lib/hello.so"))
    # # hello_lib.Hello.argtypes = [ctypes.c_char_p]
    # # hello_lib.Hello.restype = None
    # # name = "John Doe"
    # # hello_lib.Hello(name.encode('utf-8'))
    # #
    # # search_lib = ctypes.CDLL(os.path.abspath("./lib/search.so"))
    # # search_lib.searchStringInFileWrapper.argtypes = [ctypes.c_char_p, ctypes.c_char_p]
    # # search_lib.searchStringInFileWrapper.restype = ctypes.c_char_p
    # #
    # file_path = './output.log'
    # search_string = 'approvalComments'
    # result = gt.search_string_in_file_by_python(file_path, search_string)
    #
    # end_time = time.time()
    # elapsed_time = end_time - start_time
    #
    # print(f"程序总共耗时 {elapsed_time:.4f} 秒")
    # print(result)
    # print(gt.search_string_in_file_by_go(file_path, search_string))
    # print(result['Error'])
    # if result.get('Error'):
    #     print(result['Error'])
    # else:
    #     print(result['MatchedLines'])
    # c_file_path = ctypes.c_char_p(file_path.encode('utf-8'))
    # c_search_string = ctypes.c_char_p(search_string.encode('utf-8'))
    #
    # c_result = search_lib.searchStringInFileWrapper(c_file_path, c_search_string)
    # result = json.loads(ctypes.string_at(c_result).decode('utf-8'))
    # print(result)
